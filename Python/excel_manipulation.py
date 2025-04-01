
from openpyxl import load_workbook
import os
import pandas as pd

os.chdir(r"F:\Active\NFK GARO\Resources")
# Load workbook and sheet
wb = load_workbook(filename="Entity ifrån Egenskap till Nytta inkl värde i kronor.xlsx")
ws = wb.active  # or wb["SheetName"] if you know it

# Access specific cells
print("B2:", ws["B2"].value)
print("C2:", ws["C2"].value)
print("D2:", ws["D2"].value)
